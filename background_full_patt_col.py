import openpyxl
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook("students.xlsx")
ws=wb["Newsheet"]
# pattern type:
# solid, lightGray, lightGrid, darkGrid, darkUp, darkVertical, lightTrellis, lightVertical, gray125, lightHorizontal, mediumGray, darkHorizontal, gray0625, darkGray, lightUp, darkDown, lightDown, darkTrelli

pattern_full=PatternFill(patternType="solid", fgColor="FF7F50")
cell_pattern=ws['A2']
cell_pattern.fill=pattern_full
wb.save("students.xlsx")
