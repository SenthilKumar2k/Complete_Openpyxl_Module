import openpyxl

daily_report=openpyxl.load_workbook("daily_report.xlsx")
master_report=openpyxl.load_workbook("lifetime_report.xlsx")

daily=daily_report["Sheet1"]
master=master_report["Data"]

daily_purchase_count=0
daily_row_id=[]
while True:
    daily_purchase_count+=1
    data=daily.cell(row=daily_purchase_count,column=1).value
    if data==None:
        break
    daily_row_id.append(data)
print(daily_row_id)
print(daily_purchase_count)

master_purchase_count=0
master_report_id_match=[]
while True:
    master_purchase_count+=1
    id=master.cell(row=master_purchase_count,column=1).value
    if id==None:
        break
    if id in daily_row_id:
        data_list=[]
        for i in range(2,7):
            data=master.cell(row=master_purchase_count,column=i).value
            data_list.append(data)
        master_report_id_match.append(data_list)

print(master_purchase_count)
print(master_report_id_match)

for i in master.iter_rows():
    id=i[0].value
    # if id==None:
    #     break
    print(id)

for i in daily.iter_rows():
    row=i[0].row
    # if row==None:
    #     break
    print(row)
