import openpyxl

wb=openpyxl.load_workbook("students.xlsx")
ws=wb["Sheet"]
rows=ws.iter_rows(min_row=1,max_row=7, min_col=1,max_col=2) #create tuple for rows cells
print(rows)        
# it gives <generator object Worksheet._cells_by_row at 0x7fce0c6d36f0>
for row in rows:
    print(row)
#(<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>)(<Cell 'Sheet'.A2>, <Cell 'Sheet'.B2>)(<Cell 'Sheet'.A3>, <Cell 'Sheet'.B3>)
#(<Cell 'Sheet'.A4>, <Cell 'Sheet'.B4>)(<Cell 'Sheet'.A5>, <Cell 'Sheet'.B5>) etc
rows=ws.iter_rows(min_col=1,min_row=1,max_col=2,max_row=7)
for a, b in rows:
    print(a.value, b.value)
# Students  Marks
# senthil 480
# priya 450
# anu 490
# hari 470
# ammu 479
# kiran 481
rows=ws.iter_rows(min_col=1,min_row=1,max_col=2,max_row=7)
marks={}
for a, b in rows:
    marks[a.value]=b.value
print(marks)
#{'Students ': 'Marks', 'senthil': 480, 'priya': 450, 'anu': 490, 'hari': 470, 'ammu': 479, 'kiran': 481}
rows=ws.iter_rows(min_col=1,min_row=1,max_col=2,max_row=7)
student=[]
marks=[]
for a, b in rows:
    student.append(a.value)
    marks.append(b.value)
print(student,marks)
#['Students ', 'senthil', 'priya', 'anu', 'hari', 'ammu', 'kiran'] ['Marks', 480, 450, 490, 470, 479, 481]

colm=ws.iter_cols(max_row=7, min_row=1, min_col=1, max_col=2) #create tuple for columns cells
print(colm)
for col in colm:
    print(col)

rows=list(ws.rows) # rows function in sheet print all cells of row in tuple
print(rows)        # like iter_rows

colm=list(ws.columns) # columns function in sheet print all cells of column in tuple
print(colm)           # like iter_cols