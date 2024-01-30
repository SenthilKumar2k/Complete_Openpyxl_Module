import openpyxl
import re

wb=openpyxl.load_workbook("lifetime_report.xlsx")
ws=wb['Data']

row_count=1
while True:
    row_count+=1
    first=ws.cell(row=row_count, column=2).value
    if first is not None:
        print(re.sub(r'@"','',str(first)))
        ws.cell(row=row_count, column=2).value=str(first).strip('"')
    else:
        break
wb.save("lifetime_report.xlsx")