import openpyxl
from openpyxl.styles import Font

wb=openpyxl.load_workbook("lifetime_report.xlsx")
wbd=openpyxl.load_workbook("daily_report.xlsx")
master_data=wb["Data"]
daily_data=wbd["Sheet1"]

#get row count for daily_report
data_daily=True
daily_row_count=0
while data_daily:
    daily_row_count+=1
    data=daily_data.cell(row=daily_row_count, column=1).value
    if data==None:
        data_daily=False

print("row count for daily report : ",daily_row_count)

#get data from the daily_report and stored it in the list in the form of dict
daily_data_list=[]

for i in range(1,daily_row_count):
    daily_row_data={}
    daily_row_data["id"]=daily_data.cell(row=i,column=1).value
    daily_row_data["Today purchase"]=daily_data.cell(row=i,column=2).value
    daily_row_data["Today rewards"]=daily_data.cell(row=i, column=3).value
    daily_data_list.append(daily_row_data)

print(daily_data_list)
# create new xlsx file to store the newly updated data

#get row count for lifetime_report
data_master=True
master_row_count=0
while data_master:
    master_row_count+=1
    data=master_data.cell(row=master_row_count,column=1).value
    if data==None:
        data_master=False

print("row count for master report : ", master_row_count)

new_report=openpyxl.Workbook()
ws=new_report.active

data_header=True
header_col_count=0
header_col_list=[]

while data_header:
    header_col_count+=1
    header_data=master_data.cell(row=1,column=header_col_count).value
    if header_data==None:
        break
    header_col_list.append(header_data)

print(header_col_list)

header_style=Font(name="Chilanka", size=12, bold=True)

for i, header_col_name in enumerate(header_col_list):
    #print(i, header_col_name)
    ws.cell(row=1,column=i+1).value=header_col_name
    ws.cell(row=1,column=i+1).font=header_style
    # style_header=ws.cell(row=1, column=i+1)
    # style_header.font=header_style

ids=[]

for data in daily_data_list:
    ids.append(data['id'])
ids.pop(0)
print(ids)

final_data=[]
for i in range(1,master_row_count):
    id=master_data.cell(row=i, column=1).value
    if id in ids:
        exist_list=[]
        for j in range(2,7):
            exist_list.append(master_data.cell(row=i, column=j).value)
        final_data.append(exist_list)
print(id)
print(final_data)


#new_report.save("new purchase report from exist.xlsx")


