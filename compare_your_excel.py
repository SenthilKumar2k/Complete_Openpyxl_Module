import openpyxl
from openpyxl.styles import PatternFill

wb=openpyxl.load_workbook("lifetime_report.xlsx")
wbc=openpyxl.load_workbook("lifetime_report (copy).xlsx")
wb_sheet=wb['Data']
wbc_sheet=wbc['Data']

#print(wb_sheet['A5'].value)

pattern_style=PatternFill(start_color="F39C12", end_color="BB8FCE", fill_type="solid")

for row in wb_sheet.iter_rows():
    #print(row[0].value)
    for cell in row:
        if cell.value != None:
            cell_value=cell.value
            cell_location=cell.coordinate
        if cell.value==None:
            break

        if cell_value != wbc_sheet[cell_location].value:
            #print(wbc_sheet[cell_location].value)
            cell.fill=pattern_style

wb.save("compare_lifetime_report.xlsx")