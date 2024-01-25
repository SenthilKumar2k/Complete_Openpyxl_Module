import openpyxl

master_report=openpyxl.load_workbook("lifetime_report.xlsx")
daily_report=openpyxl.load_workbook("daily_report.xlsx")
master_data=master_report['Data']
daily_data=daily_report['Sheet1']

#get row count for daily_report
data_daily=True
daily_row_count=0
while data_daily:
    daily_row_count+=1
    data=daily_data.cell(row=daily_row_count, column=1).value
    if data==None:
        data_daily=False

print("row count for daily report : ",daily_row_count)

#get row count for lifetime_report
data_master=True
master_row_count=0
while data_master:
    master_row_count+=1
    data=master_data.cell(row=master_row_count,column=1).value
    if data==None:
        data_master=False

print("row count for master report : ", master_row_count)

#get data from the daily_report and stored it in the list in the form of dict
daily_data_list=[]

for i in range(1,daily_row_count):
    daily_row_data={}
    daily_row_data["id"]=daily_data.cell(row=i,column=1).value
    daily_row_data["Today purchase"]=daily_data.cell(row=i,column=2).value
    daily_row_data["Today rewards"]=daily_data.cell(row=i, column=3).value
    daily_data_list.append(daily_row_data)

print(daily_data_list)
#[{'id': 30, 'Today purchase': 1, 'Today rewards': 3}]

# get the particular row data from lifetime_report by using the id of daily report 
# add lifetime purchase and todays purchase
# add total rewards and todays rewards

for i in range(2,master_row_count):
    id=master_data.cell(row=i,column=1).value
    for row in daily_data_list:
        if row["id"]==id:
            today_purchase=row['Today purchase']
            today_reward=row['Today rewards']
            total_purchase=master_data.cell(row=i,column=5).value
            total_rewards=master_data.cell(row=i,column=6).value
            print(total_purchase,total_rewards)
            purchase=total_purchase+today_purchase
            rewards=total_rewards+today_reward
            master_data.cell(row=i,column=5).value=purchase
            master_data.cell(row=i,column=6).value=rewards

master_report.save("lifetime_report.xlsx")