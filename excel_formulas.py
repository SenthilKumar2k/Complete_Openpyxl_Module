import openpyxl
from openpyxl.styles import Font

wb=openpyxl.load_workbook("students.xlsx")
ws=wb["Sheet"]
ws["A11"]="Total"
ws["A11"].font=Font(size=15, bold=True, name="Arial")
ws['A12']='Average'
ws["B11"]='=SUM(B2:B9)'
ws["B12"]="=AVERAGE(B2:B9)"
for i in range(2,10):
    balance=ws.cell(row=i,column=2).value
    interest=ws.cell(row=i,column=3).value
    final=(balance*interest)
    final_amount=(balance*interest)+balance
    ws.cell(row=i,column=4).value=final
    ws.cell(row=i,column=5).value=final_amount
wb.save("students.xlsx")