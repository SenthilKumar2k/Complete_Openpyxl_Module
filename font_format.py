import openpyxl
from openpyxl.styles import Font, Color

wb=openpyxl.load_workbook("students.xlsx")
ws=wb['Newsheet']

fon=Font(name="Liberation Mono", size=14, color="CD5C5C", italic=True, bold=True)
block=ws["A1"]
block.font=fon

font_style=Font(name="Nimbus Mono PS", size=12, color="DB3B22", underline='single', strikethrough=True)
# underline must be one of this {'single', 'double', 'doubleAccounting', 'singleAccounting'}
for i in range(1,7):
    ws.cell(row=i, column=2).font=font_style
wb.save("students.xlsx")