import openpyxl

wb=openpyxl.load_workbook("students.xlsx")
print(wb.sheetnames)
ws=wb["Sheet1"]
print(ws)                                         #<worksheet 'Sheet1'>
print(ws['B3'])                                   #<Cell 'Sheet1'.B3>
print(ws['B3'].value)                             #display the value in the cell
print(ws.cell(row=7,column=1).value)              #display value in cell of specific row and column
value=ws['A2':'B6']
print(value)
#((<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>), (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>), 
#(<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>), (<Cell 'Sheet1'.A6>, <Cell 'Sheet1'.B6>))
for a, b in value:
    print(a.value, b.value)
#senthil 3
# priya 7
# anu 1
# hari 5
# ammu 4