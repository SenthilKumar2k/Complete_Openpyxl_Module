import openpyxl

wb=openpyxl.load_workbook("students.xlsx")
ws=wb["Sheet"]
print(ws['B5'].value)   # used to read the values in sheet cell
ws['B5'].value=471      # used to update the value of particular cell
ws['A9']="meera"        # helps to write in a particular cell
ws['B9']=453
print(ws.cell(row=6,column=2).value) # used to read the value in cell
ws.cell(row=6,column=2).value=478    # used to update the value of particular cell
ws.cell(row=1, column=3).value="Double Mark"
print(ws.cell(row=1,column=3).value) 
# Used to write or update value in multiple column cell
for i in range(2,10):
    new=ws.cell(row=i, column=2).value
    new_col=new*2
    ws.cell(row=i, column=3).value=new
wb.save("students.xlsx")
