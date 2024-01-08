import openpyxl

#Helps to create excel.xlsx file
"""workbook = openpyxl.Workbook()
sheet = workbook.active

# Add data to the sheet
sheet['A1'] = 'Hello'
sheet['B1'] = 'World!'

# Save the workbook
workbook.save('students.xlsx')
"""
try:
    wb=openpyxl.load_workbook("students.xlsx")   # Helps to load the xlsx file
    print(wb.sheetnames)                         # to display all sheet name in the xlsx file
    ws=wb["Sheet"]                               # to check which is worksheet or not
    print(ws)
    ws1=wb["Sheet2"]                             # to specify the particular worksheet
    print(ws1)
    #wb.create_sheet("Newsheet")                 # create worksheet at last
    wb.create_sheet("Sheet1",1)                  # create worksheet in the particular index
    wb.save("students.xlsx")                     # to save created sheet in xlsx file
except Exception as e:
    print("error:{}".format(e))
