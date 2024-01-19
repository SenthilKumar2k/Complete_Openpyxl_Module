import openpyxl
from openpyxl.styles import numbers

#used to change the format of value

wb=openpyxl.load_workbook("students.xlsx")
ws=wb["Sheet2"]
ws["A1"]="S.No"
ws["A1"].number_format=numbers.FORMAT_TEXT
ws["B1"]="Date"
ws["B1"].number_format=numbers.FORMAT_TEXT
ws["C1"]="Name"
ws["C1"].number_format=numbers.FORMAT_TEXT
ws["A2"]=1
ws["B2"]="20/12/24"
ws["B2"].number_format=numbers.FORMAT_DATE_DATETIME
ws["C2"]="senthil"
ws["c2"].number_format=numbers.FORMAT_TEXT
wb.save("students.xlsx")